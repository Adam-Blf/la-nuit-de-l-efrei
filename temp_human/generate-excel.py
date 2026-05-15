"""Genere un fichier Excel multi-onglets pour la campagne Insta du Prom EFREI 2026.

Onglets :
  1. Calendrier publications · 50 publications avec dates / heures / type / id / statut
  2. Captions engageantes · 15 captions premium pour les publications majeures
  3. Bios @promefrei · 6 versions avec dates de bascule
  4. Bascules bio · timeline pure des changements de bio
  5. Ressources · inventaire visuels prets / a produire

Sortie : Plan-Com-Gala-Efrei-2026.xlsx
"""

from __future__ import annotations
import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

THUMB_DIR = Path(__file__).parent / ".thumbs-excel"
THUMB_DIR.mkdir(exist_ok=True)
THUMB_W_PX = 90       # largeur cible vignette
THUMB_MAX_H_PX = 140  # hauteur max vignette
EMBED_IMAGES = False  # mis a False pour reduire la taille de l'Excel


def make_thumbnail(visual_path_str: str, root: Path, fallback_id: str | None = None) -> Path | None:
    """Resolve un visual_path CSV vers une image PNG miniature (pour embed Excel).

    Strategies de resolution (par ordre) :
    1. chemin direct .png/.jpg → utilise tel quel
    2. chemin .mp4 → cherché cover.png dans le meme dossier
    3. chemin dossier (ou se termine par /) → cherché slide-1.png, story.png, banner.png, cover.png
    4. fallback : si fallback_id fourni et chemin introuvable → tente visuels-generes/<id>/{slide-1,story,banner,cover}.png
    5. introuvable → None
    """
    s = (visual_path_str or "").strip().rstrip("/")

    def _resolve_dir(d: Path) -> Path | None:
        for cand in ("slide-1.png", "story.png", "banner.png", "cover.png"):
            if (d / cand).exists():
                return d / cand
        return None

    p = root / s if s else None

    if p is None:
        pass
    elif p.is_dir():
        p = _resolve_dir(p)
    elif s.lower().endswith(".mp4"):
        cover = p.parent / "cover.png"
        p = cover if cover.exists() else None

    # Fallback : visuels-generes/<id>/
    if (not p or not p.exists()) and fallback_id:
        gen_dir = root / "visuels-generes" / fallback_id
        if gen_dir.is_dir():
            p = _resolve_dir(gen_dir)

    if not p or not p.exists():
        return None
    if p.suffix.lower() not in (".png", ".jpg", ".jpeg"):
        return None
    s = str(p.relative_to(root)) if p.is_relative_to(root) else p.name
    # Cache thumb par hash du chemin
    safe_name = s.replace("/", "_").replace("\\", "_").replace(":", "_") + "_thumb.png"
    thumb_path = THUMB_DIR / safe_name
    if thumb_path.exists() and thumb_path.stat().st_mtime >= p.stat().st_mtime:
        return thumb_path
    try:
        img = PILImage.open(p).convert("RGB")
        img.thumbnail((THUMB_W_PX, THUMB_MAX_H_PX), PILImage.LANCZOS)
        img.save(thumb_path, "PNG", optimize=True)
        return thumb_path
    except Exception:
        return None

ROOT = Path(__file__).parent
CSV = ROOT / "posts-meta-suite.csv"
import os, time
PRIMARY = ROOT / "Plan-Com-Gala-Efrei-2026.xlsx"
# Si fichier verrouille (ouvert dans Excel), ecrit avec timestamp
try:
    if PRIMARY.exists():
        with open(PRIMARY, "ab"): pass
    OUT = PRIMARY
except (PermissionError, OSError):
    OUT = ROOT / f"Plan-Com-Gala-Efrei-2026-{time.strftime('%Y%m%d-%H%M')}.xlsx"

# Styles
NAVY = "001329"
GOLD = "B8860B"
CREAM = "F5E6D3"
WHITE = "FFFFFF"

HEADER_FILL = PatternFill(start_color=NAVY, end_color=NAVY, fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color=WHITE, size=11)
ROW_ALT_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
BORDER = Border(left=Side(style="thin", color="BBBBBB"),
                right=Side(style="thin", color="BBBBBB"),
                top=Side(style="thin", color="BBBBBB"),
                bottom=Side(style="thin", color="BBBBBB"))


def style_header_row(ws, row: int = 1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


def style_data_rows(ws, start_row: int = 2):
    for r in range(start_row, ws.max_row + 1):
        for cell in ws[r]:
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if r % 2 == 0:
                cell.fill = ROW_ALT_FILL


def autosize(ws, max_w: int = 60):
    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = 0
        for cell in col:
            if cell.value:
                lines = str(cell.value).split("\n")
                for line in lines:
                    max_len = max(max_len, len(line))
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_w)


def freeze(ws, row: int = 2):
    ws.freeze_panes = ws.cell(row=row, column=1)


FR_JOURS = ["lun", "mar", "mer", "jeu", "ven", "sam", "dim"]
FR_MOIS = ["jan", "fev", "mar", "avr", "mai", "jui", "jui", "aou", "sep", "oct", "nov", "dec"]


def date_fr(date_str: str) -> str:
    """Convertit '2026-05-11' en 'lun 11 mai 2026'."""
    from datetime import datetime
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        return f"{FR_JOURS[dt.weekday()]} {dt.day} {FR_MOIS[dt.month-1]} {dt.year}"
    except Exception:
        return date_str


def categorize(type_str: str) -> str:
    """Categorise le type de publication en emoji + label."""
    t = type_str.lower()
    if "story" in t:
        return "📱 Story"
    if "reel" in t:
        return "🎬 Reel"
    if "carrousel" in t or "image" in t:
        return "📸 Post feed"
    if "bio" in t:
        return "📝 Bio"
    if "banner" in t or "banniere" in t:
        return "🖼️ Bannière"
    if "live" in t:
        return "🔴 Live"
    return "📂 Autre"


# === ALT TEXT mapping · description accessibilite pour chaque asset ===
ALT_TEXT = {
    "post-J19-recap": "Carrousel 5 slides · récap parcours J-19 avant La Nuit de l'EFREI 2026 · hibou Barney en costume sur fond péniche nuit, texte hero or italique 'DIX-NEUF JOURS'",
    "post-J19-tript-1-le-gala": "Slide 1/3 du tryptique J-19 · 'LE GALA' grand titre italique or sur fond péniche nuit étoilée",
    "post-J19-tript-2-est": "Slide 2/3 du tryptique J-19 · 'EST' grand titre italique or sur fond intérieur péniche",
    "post-J19-tript-3-de-retour": "Slide 3/3 du tryptique J-19 · 'DE RETOUR' grand titre italique or sur fond pont péniche nuit",
    "post-J19-concours-launch": "Story teaser J-19 · annonce concours mercredi 13 mai 21h, Barney dance, fond péniche",
    "post-J18-inclusion": "Carrousel 8 slides J-18 · 'Qui peut venir ?' · tarifs diplômés 14€, étudiants/alumni 18€, écoles partenaires 18€, externes 22€",
    "story-J18-inclusion": "Story compagnon du post inclusion J-18 · question 'Qui peut venir ?' + swipe up vers le post",
    "reel-J18-inclusion": "Reel 9:16 7s · enchaînement xfade des 8 slides du post inclusion 'Qui peut venir ?'",
    "post-J16-billetterie": "Carrousel 4 slides J-16 · update billetterie 50% des places vendues, jauge progression or",
    "post-J15-pourquoi": "Carrousel 5 slides J-15 · 'Pourquoi 10 ans ?' · histoire du dernier gala 2016, retour 2026",
    "post-J15-promposition": "Carrousel 5 slides J-15 · concours PROMPOSITION officiel · 'Tu m'invites au gala ?', règles et calendrier",
    "post-J12-dresscode": "Carrousel 5 slides J-12 · 'Comment je m'habille ?' · dress code Élégant, costume robe smoking",
    "post-J10-djreveal": "Carrousel 3 slides J-10 · DJ reveal officiel · set continu minuit-04h, ambiance disco",
    "post-J9-transports": "Carrousel 5 slides J-9 · 'Comment je viens ?' · Métro 10, RER B/C, Vélib, Noctilien",
    "post-J7-recap": "Carrousel 6 slides J-7 · récap complet une semaine avant · lieu, programme, dress code, transports",
    "post-J6-dresscode": "Carrousel 1 slide J-6 · rappel dress code Élégant, sortez les housses",
    "post-J5-programme": "Carrousel 5 slides J-5 · programme officiel 22h-04h, set DJ et photobooth",
    "post-J4-timeline": "Carrousel 6 slides J-4 · 'Qu'est-ce qui se passe ?' · timeline détaillée 22h-04h",
    "post-J3-derniers": "Carrousel 1 slide J-3 · 'Derniers billets' · urgence avant sold-out",
    "post-J2-shortlist": "Carrousel 4 slides J-2 · shortlist concours PROMPOSITION · 3 finalistes en vote",
    "post-J1-demain": "Carrousel 1 slide J-1 · 'Demain' · countdown la veille du gala",
    "post-J1-gagnant": "Carrousel 1 slide J-1 · gagnant concours PROMPOSITION révélé",
    "post-J0-cesoir": "Carrousel 1 slide J-0 · 'Ce soir' · jour J 22h La Péniche",
    "post-Jplus1-bestof": "Carrousel 10 slides J+1 · best-of officiel de la nuit · 10 moments à retenir",
    "post-Jplus3-drive": "Carrousel 1 slide J+3 · Drive photos ouvert · best-of par mail",
    # Stories
    "story-J19-vestiaire": "Story J-19 countdown vestiaire · 'Pas la peine de venir léger, le vestiaire gère'",
    "story-J19-concours-push": "Story J-19 push concours · teaser mercredi 13 mai 21h",
    "story-J18-pont": "Story J-18 countdown pont supérieur · 'Le pont, c'est là que les souvenirs se prennent'",
    "story-J17-sondage": "Story J-17 countdown sondage · poll 'tu es chaud / très chaud'",
    "story-J17-concours-reminder1": "Story J-17 reminder concours · compte à rebours mercredi 21h",
    "story-J16-50pct": "Story J-16 countdown · 50% des 350 places parties",
    "story-J15-2sem": "Story J-15 countdown · 2 semaines pile",
    "story-J15-boost": "Story J-15 boost · sticker question 'tu viens avec qui ?'",
    "story-J15-nouveau-concours": "Story J-15 teaser nouveau concours mercredi 21h",
    "story-J15-pourquoi-compagnon": "Story compagnon du post J-15 'Pourquoi 10 ans ?' · CTA swipe up",
    "story-J14-dj-indice": "Story J-14 indice DJ · silhouette floutée, reveal J-10",
    "story-J14-concours-inspi": "Story J-14 inspiration concours · 'Donne-moi tes idées'",
    "story-J13-coiffeur": "Story J-13 countdown coiffeur · 'Pensez aux rdv coiffeur, barbier, manucure'",
    "story-J12-photographe": "Story J-12 countdown photographe · 'Photo et vidéo briefés'",
    "story-J12-dresscode-compagnon": "Story compagnon du post J-12 dress code Élégant · CTA swipe up",
    "story-J11-securite": "Story J-11 countdown sécurité · 'Carte étu obligatoire'",
    "story-J10-10jours": "Story J-10 countdown · 10 jours pile",
    "story-J10-concours-reminder2": "Story J-10 reminder concours · 'DM avant le 25 mai'",
    "story-J9-transports": "Story J-9 countdown transports · '7 lignes à 5 min à pied'",
    "story-J9-transports-compagnon": "Story compagnon du post J-9 transports · CTA swipe up",
    "story-J8-noctilien": "Story J-8 countdown retour de nuit · 'Noctiliens directs depuis Châtelet'",
    "story-J7-1sem": "Story J-7 countdown · 'Une semaine pile'",
    "story-J7-recap-compagnon": "Story compagnon du post J-7 récap · CTA swipe up",
    "story-J6-finalisation": "Story J-6 countdown finalisation · derniers détails programme",
    "story-J5-programme": "Story J-5 countdown programme · programme révélé",
    "story-J5-programme-compagnon": "Story compagnon du post J-5 programme · CTA swipe up",
    "story-J5-concours-48h": "Story J-5 concours · 'Plus que 48h pour ta vidéo Promposition'",
    "story-J4-tenue": "Story J-4 countdown tenue · 'Sortez la tenue. Repassez. Essayez.'",
    "story-J4-concours-demain": "Story J-4 concours · 'Deadline demain 23h59'",
    "story-J3-3jours": "Story J-3 countdown · 'Trois jours. Trois nuits. On y est presque.'",
    "story-J3-concours-1h": "Story J-3 concours rush · 'Plus qu'une heure pour DM'",
    "story-J2-48h": "Story J-2 countdown · 48h avant le gala",
    "story-J2-vote": "Story J-2 poll vote concours · sticker poll sur 24h",
    "story-J1-demain": "Story J-1 countdown · 'Demain'",
    "story-J1-concours-vote-clos": "Story J-1 votes clos · 'Verdict ce soir 19h'",
    "story-J1-gagnant-celebration": "Story J-1 gagnant célébration · 'Bravo, demain tu y es'",
    "story-J0-matin": "Story matinale J-0 11h · 'Ce soir 22h La Péniche'",
    "story-J0-jourj": "Story J-0 jour J · countdown final 22h",
    "story-J0-ouverture": "Story J-0 ouverture des portes 21h55 · 'On ouvre · La Péniche est à vous'",
    "story-J0-gagnants-onsite": "Story J-0 live · gagnants concours sur place",
    "story-Jplus1-thanks": "Story J+1 thanks · 'Merci aux 350' · Drive officiel à venir",
    "story-Jplus2-coulisses": "Story J+2 coulisses · 'L'envers du décor, merci équipe BDA'",
    "story-gagnants-concours-pod": "Story J-17 annonce gagnants concours · Melissa PHILIPPE, Nabil BENOUALI, Enora IRITZ",
    # Reels mp4
    "reel-J15-promposal": "Reel 9:16 · annonce PROMPOSITION officielle, Barney dance",
    "reel-J11-champagne": "Reel 9:16 · '28 mai La Péniche · ça va pop' · ambiance champagne",
    "reel-J7-teaser": "Reel 9:16 8s · teaser comeback une semaine avant",
    "reel-Jplus1-aftermovie": "Reel 9:16 15s · after-movie best moments de la nuit",
    # Bios
    "bascule-bio-v1": "Bascule bio Insta v1 · annonce concours vidéo, 2 places offertes",
    "bascule-bio-v2": "Bascule bio Insta v2 · J-15, concours jusqu'à J-3",
    "bascule-bio-v3": "Bascule bio Insta v3 · J-7, concours derniers jours ou sold-out",
    "bascule-bio-v4": "Bascule bio Insta v4 · DEMAIN, gagnants concours révélés",
    "bascule-bio-v5": "Bascule bio Insta v5 · post-event, 350 invités, merci",
    # Live
    "live-J0-cadence": "Stories live J-0 · cadence 1 toutes les 15-30 min de 22h à 04h",
}


def get_alt_text(asset_id: str) -> str:
    """Retourne le alt text d'un asset ou un fallback."""
    if asset_id in ALT_TEXT:
        return ALT_TEXT[asset_id]
    # Fallback generique
    if asset_id.startswith("story-"):
        return f"Story Insta · {asset_id.replace('story-', '').replace('-', ' ')}"
    if asset_id.startswith("post-"):
        return f"Post feed Insta · {asset_id.replace('post-', '').replace('-', ' ')}"
    if asset_id.startswith("reel-"):
        return f"Reel 9:16 Insta · {asset_id.replace('reel-', '').replace('-', ' ')}"
    return asset_id.replace('-', ' ')


# Map global · captions par asset id (rempli par sheet_captions au premier appel)
_CAPTIONS_BY_ID = {}


# Couleurs par categorie · fill clair pour lisibilite
CAT_COLORS = {
    "📸 Post feed":   "E8F4FA",  # bleu pale
    "📱 Story":       "FCE8F4",  # rose pale
    "🎬 Reel":        "F0E8FC",  # violet pale
    "📝 Bio":         "FCF8E8",  # jaune pale
    "🖼️ Bannière":   "E8FCEF",  # vert pale
    "🔴 Live":        "FCE8E8",  # rouge pale
    "📂 Autre":       "F0F0F0",  # gris pale
}


def _populate_calendar_sheet(ws, csv_rows, root):
    """Helper · remplit une feuille de calendrier avec colonnes standardisees + couleurs categorie + data validation case a cocher."""
    ws.append([
        "Postee?", "Date FR", "Heure", "Categorie", "Type detail",
        "ID", "Statut", "Bio active", "Description (caption)",
        "Alt text (accessibilite)", "Reference fiche", "Visuel path", "Apercu"
    ])
    style_header_row(ws)
    rows_with_visual = []
    for row in csv_rows:
        if len(row) < 8:
            row = row + [""] * (8 - len(row))
        date_iso, heure, plateforme, type_str, vid, statut, fiche, path = row
        # Description · caption complete si dispo, sinon alt_text comme description courte
        caption = _CAPTIONS_BY_ID.get(vid, "")
        description = caption if caption else get_alt_text(vid)
        ws.append([
            "☐",
            date_fr(date_iso), heure, categorize(type_str), type_str,
            vid, statut, bio_active_at(date_iso, heure) if heure else "",
            description,
            get_alt_text(vid), fiche, path, ""
        ])
        thumb = make_thumbnail(path, root, fallback_id=vid)
        if thumb:
            rows_with_visual.append((ws.max_row, thumb))
    style_data_rows(ws)
    autosize(ws, max_w=55)
    freeze(ws)
    ws.row_dimensions[1].height = 32

    # Apres suppression de Date ISO · 13 colonnes au lieu de 14
    widths = {
        "A": 9,   # Postee?
        "B": 18,  # Date FR
        "C": 7,   # Heure
        "D": 16,  # Categorie
        "E": 16,  # Type detail
        "F": 32,  # ID
        "G": 22,  # Statut
        "H": 28,  # Bio active
        "I": 60,  # Description
        "J": 55,  # Alt text
        "K": 18,  # Fiche ref
        "L": 35,  # Path
        "M": 14,  # Apercu
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    from openpyxl.styles import Alignment as Align
    from openpyxl.styles import PatternFill as Fill
    from openpyxl.worksheet.datavalidation import DataValidation

    # Data validation pour case a cocher · liste deroulante ☐ / ☑
    dv = DataValidation(type="list", formula1='"☐,☑"', allow_blank=True)
    dv.add(f"A2:A{ws.max_row}")
    ws.add_data_validation(dv)

    for r in range(2, ws.max_row + 1):
        ws.cell(row=r, column=1).alignment = Align(horizontal="center", vertical="center")
        # Description et Alt text en wrap text pour visibilite
        ws.cell(row=r, column=9).alignment = Align(vertical="top", wrap_text=True)
        ws.cell(row=r, column=10).alignment = Align(vertical="top", wrap_text=True)
        # Couleur de fond selon la categorie
        cat = ws.cell(row=r, column=4).value
        color = CAT_COLORS.get(cat, None)
        if color:
            fill = Fill(start_color=color, end_color=color, fill_type="solid")
            for col in range(1, 14):
                ws.cell(row=r, column=col).fill = fill

    # Embed thumbnails dans col M (13eme) · skip si EMBED_IMAGES=False
    if EMBED_IMAGES:
        for excel_row, thumb_path in rows_with_visual:
            try:
                xl_img = XLImage(str(thumb_path))
                with PILImage.open(thumb_path) as pim:
                    tw, th = pim.size
                xl_img.width = tw
                xl_img.height = th
                cell = ws.cell(row=excel_row, column=13)
                ws.add_image(xl_img, cell.coordinate)
                ws.row_dimensions[excel_row].height = max(ws.row_dimensions[excel_row].height or 0, th * 0.78)
            except Exception as e:
                print(f"  WARN ligne {excel_row} thumbnail KO : {e}")


# === Onglet 1 : Calendrier publications COMPLET ===
def sheet_calendrier(wb: Workbook):
    ws = wb.create_sheet("Calendrier complet")
    csv_rows = []
    if CSV.exists():
        with CSV.open(encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)
            csv_rows = list(reader)
    _populate_calendar_sheet(ws, csv_rows, ROOT)


# === Feuilles separees par categorie ===
def sheets_par_categorie(wb: Workbook):
    """Cree 5 feuilles · Posts feed · Stories · Reels · Bios · Bannieres."""
    if not CSV.exists():
        return
    with CSV.open(encoding="utf-8") as f:
        reader = csv.reader(f)
        next(reader)
        all_rows = list(reader)

    # Grouper par categorie (sur le type)
    groups = {
        "📸 Posts feed": [],
        "📱 Stories": [],
        "🎬 Reels": [],
        "📝 Bios": [],
        "🖼️ Bannieres": [],
    }
    for row in all_rows:
        if len(row) < 4:
            continue
        t = row[3].lower()
        if "story" in t or "stories" in t or "story poll" in t:
            groups["📱 Stories"].append(row)
        elif "reel" in t:
            groups["🎬 Reels"].append(row)
        elif "carrousel" in t or "image" in t:
            groups["📸 Posts feed"].append(row)
        elif "bio" in t:
            groups["📝 Bios"].append(row)
        elif "banner" in t or "banniere" in t:
            groups["🖼️ Bannieres"].append(row)

    for sheet_name, rows in groups.items():
        if not rows:
            continue
        # Nom sheet : retire emoji pour compat Excel
        clean_name = sheet_name.split(" ", 1)[1] if " " in sheet_name else sheet_name
        ws = wb.create_sheet(clean_name)
        _populate_calendar_sheet(ws, rows, ROOT)


# === Onglet 1 : Calendrier publications (OBSOLETE - DUPLICATE - REMOVED below) ===
def _OBSOLETE_sheet_calendrier(wb: Workbook):
    ws = wb.create_sheet("OBSOLETE")
    ws.append([
        "Date FR", "Date ISO", "Heure", "Categorie", "Type detail", "ID",
        "Statut", "Bio active", "Alt text (accessibilite)", "Reference fiche",
        "Visuel path", "Apercu"
    ])
    style_header_row(ws)
    rows_with_visual = []
    if CSV.exists():
        with CSV.open(encoding="utf-8") as f:
            reader = csv.reader(f)
            next(reader)
            for row in reader:
                # row CSV : [date, heure, plateforme, type, id, statut, fiche, path]
                if len(row) < 8:
                    row = row + [""] * (8 - len(row))
                date_iso, heure, plateforme, type_str, vid, statut, fiche, path = row
                date_fr_str = date_fr(date_iso)
                cat = categorize(type_str)
                bio = bio_active_at(date_iso, heure) if heure else ""
                alt = get_alt_text(vid)
                ws.append([
                    date_fr_str, date_iso, heure, cat, type_str, vid,
                    statut, bio, alt, fiche, path, ""
                ])
                thumb = make_thumbnail(path, ROOT, fallback_id=vid)
                if thumb:
                    rows_with_visual.append((ws.max_row, thumb))
    style_data_rows(ws)
    autosize(ws, max_w=55)
    freeze(ws)
    ws.row_dimensions[1].height = 32

    # Largeurs colonnes specifiques
    ws.column_dimensions["A"].width = 18  # Date FR
    ws.column_dimensions["B"].width = 12  # Date ISO
    ws.column_dimensions["C"].width = 7   # Heure
    ws.column_dimensions["D"].width = 16  # Categorie
    ws.column_dimensions["E"].width = 16  # Type detail
    ws.column_dimensions["F"].width = 32  # ID
    ws.column_dimensions["G"].width = 22  # Statut
    ws.column_dimensions["H"].width = 28  # Bio active
    ws.column_dimensions["I"].width = 55  # Alt text
    ws.column_dimensions["J"].width = 18  # Fiche ref
    ws.column_dimensions["K"].width = 35  # Path
    ws.column_dimensions["L"].width = 14  # Apercu

    # Embed thumbnails dans col L (12eme col)
    for excel_row, thumb_path in rows_with_visual:
        try:
            xl_img = XLImage(str(thumb_path))
            with PILImage.open(thumb_path) as pim:
                tw, th = pim.size
            xl_img.width = tw
            xl_img.height = th
            cell = ws.cell(row=excel_row, column=12)
            ws.add_image(xl_img, cell.coordinate)
            ws.row_dimensions[excel_row].height = max(ws.row_dimensions[excel_row].height or 0, th * 0.78)
        except Exception as e:
            print(f"  WARN ligne {excel_row} thumbnail KO : {e}")


# === Onglet 2 : Publications · Post · Date · J-X · Type · Horaire · Description · Bio active ===
def bio_active_at(date_str: str, time_str: str) -> str:
    """Retourne la version de bio active a un instant donne."""
    from datetime import datetime
    dt = datetime.strptime(f"{date_str} {time_str}", "%Y-%m-%d %H:%M")
    bascules = [
        (datetime(2026, 5, 11, 9, 0), "v1 · concours video · 2 places offertes"),
        (datetime(2026, 5, 13, 19, 0), "v2 · J-15 · concours jusqu'a J-3"),
        (datetime(2026, 5, 21, 13, 0), "v3 ou v3-bis · J-7 · concours derniers jours / sold-out"),
        (datetime(2026, 5, 27, 19, 0), "v4 · DEMAIN · gagnants concours révélés"),
        (datetime(2026, 5, 29, 12, 0), "v5 · post-event · 350 invités · merci"),
    ]
    current = "v0 · Association pour le gala de l'EFREI"
    for bdt, label in bascules:
        if dt >= bdt:
            current = label
        else:
            break
    return current


def sheet_captions(wb: Workbook):
    ws = wb.create_sheet("Publications", 0)
    ws.append(["Post", "Date", "J-X", "Type", "Horaire", "Description", "Bio active"])
    style_header_row(ws)

    captions = [
        # (post_id, date, J-X, type, horaire, description)
        # ORDRE INVERSE pour que la grille profil affiche correctement « LE GALA | EST | DE RETOUR »
        # Insta met les + recents en haut a gauche, donc on poste DE RETOUR en 1er (10h) et LE GALA en dernier (18h)
        ("post-J19-tript-3-de-retour", "2026-05-09", "J-19", "Carrousel 4:5", "10:00",
         "🎩 DE RETOUR. (1/3)\n\nLe gala est de retour. Mais je vous laisse découvrir l'énoncé dans l'ordre · regarde les 2 prochains posts ou viens sur mon profil dans la journée.\n\nLien en bio\n\n#LaNuitDeLEFREI #PromEfrei2026 #10ansapres #EfreiParis #LaPeniche #BDAEfrei #SeineEtoilee"),
        ("post-J19-tript-2-est", "2026-05-09", "J-19", "Carrousel 4:5", "14:00",
         "✨ EST. (2/3)\n\nDeuxième partie du tryptique.\nLa suite à 18h.\nMais reviens sur le profil pour voir l'image complète.\n\nLien en bio\n\n#LaNuitDeLEFREI #PromEfrei2026 #10ansapres #EfreiParis #LaPeniche #BDAEfrei #SeineEtoilee"),
        ("post-J19-tript-1-le-gala", "2026-05-09", "J-19", "Carrousel 4:5", "18:00",
         "🌙 LE GALA. (3/3)\n\nMaintenant regarde le profil. Ça se lit · LE GALA EST DE RETOUR. ✨\n\n28 mai 2026. La Péniche. 22h. 350 places. Pas une de plus.\n10 ans plus tard, on remet ça.\n\nLien en bio\n\n#LaNuitDeLEFREI #PromEfrei2026 #10ansapres #EfreiParis #LaPeniche #BDAEfrei #SeineEtoilee"),
        ("post-J19-recap", "2026-05-09", "J-19", "Carrousel 5", "12:00",
         "🍾 10 ans après, on ramène le gala.\n\n28 mai, La Péniche, 350 places. Tout ce qu'il faut savoir avant que ce soit complet.\n\nLien en bio · HelloAsso.\n\n"),
        ("post-J19-concours-launch", "2026-05-09", "J-19", "Carrousel 4:5", "20:00",
         "🎁 Mercredi 13 mai · 21h.\n\nOn prépare un truc. 2 places à gagner à la clé.\n\nRDV mercredi soir, ça vaut le coup d'y être.\n\nLien en bio.\n\n"),
        ("post-J18-inclusion", "2026-05-10", "J-18", "Carrousel 4:5", "13:00",
         "🎩 Le gala c'est pour qui ?\n\nPour tout le monde. Diplômés EFREI 2025 à 14€ (cadeau de fin), étudiants EFREI / alumni / Groupe Assas à 18€, écoles partenaires aussi à 18€, externes à 22€.\n\nTa place inclut 2 consos à l'embarquement (mousseux ou cocktail de vin + soft) + petits fours.\n\n350 places, une nuit. Pense à réserver tant qu'il en reste.\n\nLien en bio · HelloAsso.\n\n#LaNuitDeLEFREI #PromEfrei2026"),
        ("reel-champagne", "2026-05-11", "J-17", "Reel 9:16", "20:00",
         "28 mai. 22h. La Péniche. Et ça va pop. 🍾✨\n\n(C'est du Champomy, on vous rassure 😉\nMais le 28 au soir, on vous laisse choisir.)\n\nLien en bio\n\n"),
        ("post-J16-billetterie", "2026-05-12", "J-16", "Carrousel 4", "12:30",
         "🎟️ Moitié des places parties. ⚡\nPlus que 175 billets avant que ce truc soit ferme.\nSi tu sens que c'est pour toi, c'est maintenant.\n\nLien en bio\n\n"),
        ("post-J15-promposition", "2026-05-13", "J-15", "Carrousel 4:5", "20:00",
         "💌 La PROMPOSITION officielle.\n\nLe concours de la plus belle invitation de cavalier(e) pour le gala.\n\nComment ·\n1. Tu te filmes en train d'inviter quelqu'un au gala (formule libre · surprise, sketch, ce que tu veux)\n2. DM @promefrei avant lundi 25 mai 23h59\n3. Top 3 sélectionné mardi 26 mai, vote public en story\n4. Mercredi 27 mai 19h, on annonce le gagnant\n5. Vous y allez tous les deux gratuit, billets remboursés après le gala\n\nGet creative.\n\nLien en bio\n\n"),
        ("reel-promposal", "2026-05-13", "J-15", "Reel 9:16", "21:00",
         "💌 « Tu m'invites au gala ? »\n\nLa PROMPOSITION est lancee.\nLe concours de la plus belle invitation de cavalier(e). 2 places à gagner.\n\nVideo en DM @promefrei avant lundi 25 mai 23h59.\nLes règles dans le post juste avant.\n\nC'est la PROMPOSITION qu'on attendait.\n\nLien en bio\n\n"),
        ("post-J10-djreveal", "2026-05-18", "J-10", "Carrousel 3", "20:00",
         "🎧 Vous aviez peur que ce soit naze ? Non.\nOn a cherché le meilleur et on l'a trouvé.\nCe soir on vous présente officiellement le DJ de La Nuit de l'EFREI. 🎶\nEt oui, ça va cartonner.\n\nLien en bio\n\n"),
        ("post-J7-recap", "2026-05-21", "J-7", "Carrousel 6", "13:00",
         "⏳ Une semaine.\n\nTu te demandes ce qui t'attend ? Slide après slide, on récap tout · quand, où, comment ça marche, et pourquoi ça vaut le coup.\n\nLien en bio.\n\n"),
        ("post-J16-dresscode", "2026-05-12", "J-16", "Carrousel 5", "13:00",
         "👗🤵 Le dress code.\n\nThème · élégant. Sors ta plus belle tenue · costume, robe longue, smoking, ce que tu veux du moment que t'y mets le paquet.\n\nLe 28 mai à minuit sur le pont sup', t'as intérêt à avoir la classe.\n\nLien en bio.\n\n#LaNuitDeLEFREI #DressCode"),
        ("post-J6-dresscode", "2026-05-22", "J-6", "Carrousel 4:5", "13:00",
         "Rappel dress code · élégant pour mercredi 28.\n\nSi t'as pas encore choisi ta tenue, sors la housse ce week-end. Plus que 6 jours.\n\nLien en bio.\n\n#LaNuitDeLEFREI #DressCode"),
        ("post-J15-pourquoi", "2026-05-13", "J-15", "Carrousel 5", "13:00",
         "🌙 Pourquoi 10 ans ?\n\nLe dernier gala remonte à 2016. Depuis, 10 promos sont passées sans avoir leur soirée de fin d'études.\n\nEn 2026 on rouvre. Une nuit à quai sous Notre-Dame, 10 ans plus tard.\n\nLien en bio · HelloAsso.\n\n#LaNuitDeLEFREI #10ansapres"),
        ("post-J9-transports", "2026-05-19", "J-9", "Carrousel 5", "13:00",
         "🚇 Comment tu viens à La Péniche le 28 mai.\n\nLe plus simple · métro 10 Maubert-Mutualité, 5 min à pied. Sinon RER B ou C arrêt Saint-Michel-Notre-Dame à 7 min, des Vélib à 200m, ou un Uber direct quai de la Tournelle.\n\nRetour de nuit · les Noctiliens N12, N15 et N122. Pense à check ton itinéraire avant de finir au pied de la péniche sans plan B.\n\nLien en bio.\n\n#LaNuitDeLEFREI #LaPeniche"),
        ("post-J4-timeline", "2026-05-24", "J-4", "Carrousel 6", "13:00",
         "🌙 Le déroulé du 28 mai.\n\n22h · ouverture sur le pont sup' avec un set acoustique et un verre d'arrivée. À partir de 23h on bascule sur le bar, networking, petits fours. Minuit · tout le monde sur le pont, on est sous Notre-Dame. De minuit à 4h · DJ set + photobooth qui tourne toute la nuit.\n\n6h, ça passe vite.\n\nLien en bio.\n\n#LaNuitDeLEFREI"),
        ("post-Jplus3-drive", "2026-05-31", "J+3", "Image 1:1", "19:00",
         "📁 Drive officiel ouvert.\n\nLes photos · le reel after-movie · le best-of de la nuit.\nTout dans ta boîte mail · alerte par push.\n\nMerci aux 350.\nVivement 2027.\n\nLien en bio\n\n#LaNuitDeLEFREI #2026"),
        ("reel-J7-teaser", "2026-05-21", "J-7", "Reel 9:16", "20:00",
         "🎬 8 secondes.\nLa Seine. Les lumières. Une nuit qu'on ne racontera pas, qu'on va vivre.\n\nLien en bio\n\n"),
        ("post-J5-programme", "2026-05-23", "J-5", "Carrousel 5", "13:00",
         "🌙 22h à 4h du matin.\nDJ, champagne, photobooth, et des gens qu'on ne s'attend pas à voir là.\nVoici la nuit slide par slide.\n\nLien en bio\n\n"),
        ("post-J3-derniers", "2026-05-25", "J-3", "Carrousel 4:5", "13:00",
         "⏳ Il en reste vraiment plus beaucoup.\n\nÇa va être complet avant le week-end. Si t'hésitais, c'est le moment.\n\nLien en bio.\n\n"),
        ("post-J2-shortlist", "2026-05-26", "J-2", "Carrousel 4", "18:00",
         "🏆 Vous avez voté. Voici les 3 finalistes du concours video.\n24h pour choisir qui mérite les 2 places gratuites.\nLes votes commencent maintenant, sticker poll dans la story juste après.\n\nLien en bio\n\n"),
        ("post-J1-demain", "2026-05-27", "J-1", "Carrousel 4:5", "13:00",
         "🔥 Demain.\nC'est maintenant qu'on y pense.\nDemain soir tout le monde va parler de ca.\nTu es ou toi ?\n\nLien en bio\n\n"),
        ("post-J1-gagnant", "2026-05-27", "J-1", "Carrousel 4:5", "19:00",
         "🏆 Et le gagnant du concours « Tu m'invites au gala ? » est...\n[NOM GAGNANT].\nDemain soir, ils amènent qui ils veulent, gratuit, et la video passe en story officiel.\nBravo.\n\nLien en bio\n\n"),
        ("post-J0-cesoir", "2026-05-28", "J-0", "Carrousel 4:5", "13:00",
         "🌙 Ce soir.\n\n22h, La Péniche, on y est. Costume sorti, chaussures faites, cœur qui commence à s'emballer.\n\nÀ tout de suite.\n\nLien en bio.\n\n"),
        ("post-Jplus1-bestof", "2026-05-29", "J+1", "Carrousel 10", "12:00",
         "✨ La Nuit de l'EFREI, c'était fou.\nVoici les 10 moments qu'on ne va pas oublier.\nMerci aux 350, au BDA, à la Péniche, au DJ, au photographe.\nLe Drive complet · à venir par mail.\n\nLien en bio\n\n"),
        # ── 8 nouveaux posts feed (1 par jour pour couvrir tout le calendrier) ──
        ("post-J17-promposition", "2026-05-11", "J-17", "Image 4:5", "13:00",
         "💌 Promposition.\n\nT'as envie d'inviter quelqu'un mais tu manques de courage ? Fais ta demande en story et mentionne @promefrei. La plus belle vidéo repart avec 2 places.\n\nReveal mercredi 21h.\n\nLien en bio · HelloAsso.\n\n#LaNuitDeLEFREI #Promposition"),
        ("post-J14-dj-mystere", "2026-05-14", "J-14", "Image 4:5", "13:00",
         "🎧 Le DJ.\n\nIl a déjà mixé au Rex et il va tenir 4h chez nous. Pour le reste, faut attendre vendredi 21h.\n\nDis-nous en commentaire quels sons tu veux entendre, on transmet (en partie).\n\nLien en bio.\n\n#LaNuitDeLEFREI #DJReveal"),
        ("post-J13-prep", "2026-05-15", "J-13", "Image 4:5", "13:00",
         "✨ Plus que 13 jours.\n\nC'est le moment de prendre rdv chez le coiffeur, l'esthéticienne, sortir la housse de la tenue. Le 28 au soir tu vas pas avoir le temps.\n\nDress code · élégant.\n\nLien en bio · HelloAsso.\n\n#LaNuitDeLEFREI"),
        ("post-J12-photographe", "2026-05-16", "J-12", "Image 4:5", "13:00",
         "📸 Les photos · @efreipicturestudio toute la nuit.\n\nPhotobooth sur le pont sup' + photographes EPS qui tournent un peu partout. Toutes les photos vous arrivent par mail après.\n\nVous allez avoir la classe.\n\nLien en bio · HelloAsso.\n\n#LaNuitDeLEFREI #EfreiPictureStudio"),
        ("post-J11-securite", "2026-05-17", "J-11", "Image 4:5", "13:00",
         "🪪 Pièce d'identité à l'entrée.\n\nCarte étudiante + CNI ou passeport. Sans ID tu rentres pas, on peut rien faire.\n\nVérifie ce soir, pas la veille.\n\nLien en bio · HelloAsso.\n\n#LaNuitDeLEFREI"),
        ("post-J8-noctilien", "2026-05-20", "J-8", "Image 4:5", "13:00",
         "🌙 Pour rentrer le soir.\n\nLes Noctiliens N12, N15 et N122 passent à Saint-Michel et Notre-Dame, à 5 min de la péniche. Dernier passage 4h pile.\n\nCheck ton itinéraire avant de venir, pas après le DJ set.\n\nLien en bio.\n\n#LaNuitDeLEFREI"),
        ("post-Jplus2-coulisses", "2026-05-30", "J+2", "Image 4:5", "13:00",
         "🙏 Merci à toute l'équipe.\n\nLe BDA, prom'EFREI, La Péniche, le DJ, EPS, les bénévoles. Un an de prep pour qu'on puisse vivre cette nuit-là. Sans vous il s'est rien passé.\n\nLe drive arrive bientôt par mail.\n\nLien en bio.\n\n#LaNuitDeLEFREI"),
        ("post-Jplus5-rembours", "2026-06-02", "J+5", "Image 4:5", "13:00",
         "🏆 Promposition · remboursement fait.\n\nLes 2 places du couple gagnant viennent d'être remboursées. Bravo à eux, et merci à tous ceux qui ont envoyé une vidéo.\n\nLe drive officiel est ouvert, vos photos sont dedans.\n\nLien en bio.\n\n#LaNuitDeLEFREI"),
        # Post officiel concours Promposition · mercredi 13 mai 18h
        ("post-J15-promposition", "2026-05-12", "J-16", "Carrousel 5", "18:30",
         "💌 PROMPOSITION · le concours.\n\nT'as quelqu'un à inviter au gala ? On veut voir ça.\n\nLe principe · t'as déjà ta place sur HelloAsso, tu nous envoies ta vidéo d'invitation en DM @promefrei. Fais comme tu le sens · une surprise, une pancarte, un sketch qui sent un peu le désespoir, on prend tout.\n\nOn garde nos 3 préférées, vous votez en story, le couple gagnant repart avec 2 places remboursées après le gala.\n\n📅 Shortlist · mardi 26 mai.\n\nLien en bio · HelloAsso.\n\n#LaNuitDeLEFREI #Promposition #PromEfrei2026"),
    ]
    def to_fr(h: str) -> str:
        # Format FR : "18:00" -> "18h00"  ·  "9:00" -> "9h00"
        hh, mm = h.split(":")
        return f"{int(hh)}h{mm}"

    for c in captions:
        post_id, date, jx, typ, horaire, desc = c[0], c[1], c[2], c[3], c[4], c[5]
        # Populate global map pour reutilisation dans les autres feuilles (Calendrier complet, Posts feed, etc.)
        _CAPTIONS_BY_ID[post_id] = desc
        bio = bio_active_at(date, horaire)
        horaire_fr = to_fr(horaire)
        # Date au format FR (lun 11 mai 2026) au lieu de ISO
        ws.append([post_id, date_fr(date), jx, typ, horaire_fr, desc, bio])
    style_data_rows(ws)
    ws.column_dimensions["A"].width = 26  # Post
    ws.column_dimensions["B"].width = 12  # Date
    ws.column_dimensions["C"].width = 8   # J-X
    ws.column_dimensions["D"].width = 14  # Type
    ws.column_dimensions["E"].width = 10  # Horaire
    ws.column_dimensions["F"].width = 70  # Description
    ws.column_dimensions["G"].width = 40  # Bio active
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 180
    ws.row_dimensions[1].height = 32
    freeze(ws)


# === Onglet 3 : Bios @promefrei ===
def sheet_bios(wb: Workbook):
    ws = wb.create_sheet("Bios promefrei")
    ws.append(["Version", "Bascule (date + heure)", "Bio (max 150 chars)", "Lien web bio", "Phase"])
    style_header_row(ws)

    bios = [
        ("v0", "actuelle (depuis creation)", "👑 Association pour le gala de l'EFREI", "HelloAsso", "Avant lancement campagne · J-60 a J-19"),
        ("v1", "lun 11 mai 09h00", "👑 La Nuit de l'EFREI · 28.05.26 🚢\n🎬 Concours video · 2 places offertes\n🎟️ Billetterie ↓", "HelloAsso billetterie", "Lancement concours · J-17 a J-15"),
        ("v2", "mer 13 mai 19h00", "👑 La Nuit de l'EFREI · J-15 🌟\n🎬 Concours video · DM jusqu'a J-3\n🎟️ Billetterie ↓", "HelloAsso billetterie", "Mid-campaign · J-15 a J-7"),
        ("v3", "jeu 21 mai 13h00 (places dispo)", "👑 La Nuit de l'EFREI · J-7 ⚡\n🎬 Concours · videos jusqu'a lun 23h59\n🎟️ Place ↓ HelloAsso", "HelloAsso billetterie", "Phase finale · J-7 a J-1"),
        ("v3-bis", "jeu 21 mai 13h00 (sold-out)", "👑 La Nuit de l'EFREI · J-7 ✅\n🥂 Sold out · 350/350 · merci\n🌐 Programme + acces ↓", "prom.efrei.fr", "Si SOLD-OUT confirme · bascule lien"),
        ("v4", "mer 27 mai 19h00", "👑 La Nuit de l'EFREI · DEMAIN 🌟\n🥇 Gagnants concours révélés\n📍 La Péniche · 22h · 28.05 ↓", "prom.efrei.fr", "Veille du gala · J-1 a J-0"),
        ("v5", "ven 29 mai 12h00", "👑 La Nuit de l'EFREI · 28.05.26 ✨\n🌟 350 invités · 10 ans rattrapés · merci\n📁 Drive officiel ↓", "prom.efrei.fr", "Post-event permanent · J+1 a inf"),
    ]
    for b in bios:
        ws.append(list(b))
    style_data_rows(ws)
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 35
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 80
    ws.row_dimensions[1].height = 32
    freeze(ws)


# === Onglet 4 : Bascules bio (timeline pure) ===
def sheet_timeline(wb: Workbook):
    ws = wb.create_sheet("Timeline bascules bio")
    ws.append(["Date", "Heure", "Action", "Detail"])
    style_header_row(ws)
    timeline = [
        ("2026-05-11", "09:00", "Bascule bio v1", "Depuis app mobile Insta · v1 lancement concours"),
        ("2026-05-13", "19:00", "Bascule bio v2", "Depuis app mobile Insta · v2 J-15 mid-campaign"),
        ("2026-05-21", "13:00", "Bascule bio v3 ou v3-bis", "v3 si encore des places · v3-bis si SOLD-OUT (changer aussi le lien web HelloAsso → prom.efrei.fr)"),
        ("2026-05-27", "19:00", "Bascule bio v4", "Depuis app mobile · v4 DEMAIN · changer lien web sur prom.efrei.fr"),
        ("2026-05-29", "12:00", "Bascule bio v5", "Depuis app mobile · v5 post-event permanent"),
    ]
    for t in timeline:
        ws.append(list(t))
    style_data_rows(ws)
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 80
    ws.row_dimensions[1].height = 32
    freeze(ws)


# === Onglet 5 : Ressources ===
def sheet_ressources(wb: Workbook):
    ws = wb.create_sheet("Ressources visuels")
    ws.append(["Categorie", "Asset", "Format", "Statut", "Path"])
    style_header_row(ws)
    items = [
        ("Logos", "Prom Efrei (cerclé hibou)", "SVG", "OK", "logos/prom-efrei.svg"),
        ("Logos", "Prom Efrei raster", "PNG", "OK", "logos/prom-efrei-raster.png"),
        ("Logos", "EFREI institutionnel blanc", "PNG", "OK", "logos/efrei-blanc.png"),
        ("Logos", "EFREI institutionnel classique", "SVG", "OK", "logos/efrei-classique.svg"),
        ("Logos", "EFREI institutionnel noir", "PNG", "OK", "logos/efrei-noir.png"),
        ("Logos", "BDA Efrei horizontal", "SVG + PNG", "OK", "logos/bda_logo_horizontal.*"),
        ("Logos", "Mascotte Barney", "PNG", "OK", "logos/barney-mascotte.png"),
        ("Visuels statiques", "30 PNG generes (12 publications)", "PNG 1080×1350 / 1080×1080 / 1080×1920", "OK", "visuels-generes/"),
        ("Stories countdown", "31 PNG legacy J-30 a J-0", "PNG 1080×1920", "OK", "visuels-stories-pretes/"),
        ("Reel champagne", "Vidéo finale avec overlay logo", "MP4 576×1024 17s", "OK · publi 11 mai", "tiktok-champagne/reel-J11-champagne-FINAL.mp4"),
        ("Reel promposal", "A tourner mar 12 mai, publi mer 13 mai 21h", "MP4 a produire", "TODO", "tiktok-promposal/ (a creer)"),
        ("Vidéos countdown", "20 transitions split-flap J-20 a J-0", "MP4 1080×1920 4s", "OK", "countdown-video/transition-*.mp4"),
        ("Photos La Péniche", "12 photos officielles", "JPG/WEBP", "OK", "Documents 00_EFREI_Gala_2026/PromEfrei.../08_Photos_Peniche/"),
        ("Polices", "18 fontes locales (Fraunces, Cinzel, Cormorant, Inter, etc.)", "TTF", "OK", "Gala_workspace/communication/fonts-tmp/"),
    ]
    for i in items:
        ws.append(list(i))
    style_data_rows(ws)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 50
    ws.row_dimensions[1].height = 32
    freeze(ws)


# === Build ===
def main():
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)
    # Onglet 1 = Publications (vue principale demandee par Adam · Post Date J-X Type Horaire Description)
    sheet_captions(wb)
    sheet_calendrier(wb)
    sheets_par_categorie(wb)   # nouvelles feuilles par type (Posts feed, Stories, Reels, Bios, Bannieres)
    sheet_bios(wb)
    sheet_timeline(wb)
    sheet_ressources(wb)
    wb.save(OUT)
    print(f"OK · {OUT}")
    print(f"Onglets : {wb.sheetnames}")


if __name__ == "__main__":
    main()
